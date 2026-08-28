"""
Tests para la lógica del dashboard y el export (Excel/PDF) de resumen diario.

No requieren MySQL: se mockea get_db_connection() con cursores que devuelven
datos sintéticos, verificando el procesamiento de los resultados.

Orden de fetchone/fetchall en dashboard_metricas (caché vacío, confirmado_at
NO existe → el chequeo de listo_at se salta por short-circuit y no consume
fetchone):
  fetchone: [SHOW_COLUMNS confirmado_at, pedidos_hoy, ventas, producto_top,
             mesas, mesa_top, SHOW_TABLES cierres_mesa, categoria_top_row,
             total_items_row (solo si categoria_top_row no es None)]
  fetchall: [estados, cobros_metodos (solo si cierres_mesa existe), estados_hoy]

Si confirmado_at SÍ existe, se suma un fetchone extra (SHOW_COLUMNS listo_at)
antes del fetchone de tiempo_row (solo si listo_at también existe).

Orden en resumen_hoy_excel (caché vacío):
  fetchone: [SHOW_COLUMNS, resumen, mesa_top, SHOW_TABLES]
  fetchall: [por_estado, productos_top, por_hora]
"""
import io
import pytest
from datetime import date, timedelta
from unittest.mock import MagicMock, patch


# ────────────────────────────────────────────────────────────────────────────
# Fixture — limpia el caché entre tests
# ────────────────────────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def reset_col_cache():
    """
    _col_cache es global en reportes.py. Si queda poblado de un test anterior,
    las llamadas SHOW COLUMNS / SHOW TABLES se saltan y el orden del mock cambia.
    Lo limpiamos antes y después de cada test.
    """
    from app.routes.reportes import _col_cache
    _col_cache.clear()
    yield
    _col_cache.clear()


# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

def _mock_conn(fetchone_seq, fetchall_seq):
    cursor = MagicMock()
    cursor.fetchone.side_effect = list(fetchone_seq)
    cursor.fetchall.side_effect = list(fetchall_seq)
    conn = MagicMock()
    conn.cursor.return_value = cursor
    return conn


def _call_dashboard(conn):
    from app.routes.reportes import dashboard_metricas
    with patch("app.routes.reportes.get_db_connection", return_value=conn), \
         patch("app.routes.reportes.close_db_connection"):
        return dashboard_metricas(current_user={"id": 1, "rol": "admin"})


def _call_resumen_hoy(conn, fecha=None, formato="excel"):
    from app.routes.reportes import resumen_hoy_excel
    with patch("app.routes.reportes.get_db_connection", return_value=conn), \
         patch("app.routes.reportes.close_db_connection"):
        return resumen_hoy_excel(
            fecha=fecha, formato=formato, current_user={"id": 1, "rol": "admin"}
        )


def _texto_workbook(response) -> str:
    """Concatena el texto de todas las celdas de todas las hojas del xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.body))
    partes = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            partes.extend(str(c) for c in row if c is not None)
    return "\n".join(partes)


def _call_ventas_semana(conn):
    from app.routes.reportes import ventas_ultima_semana
    with patch("app.routes.reportes.get_db_connection", return_value=conn), \
         patch("app.routes.reportes.close_db_connection"):
        return ventas_ultima_semana(current_user={"id": 1, "rol": "admin"})


# ────────────────────────────────────────────────────────────────────────────
# /reportes/dashboard — campo mesa_top
# ────────────────────────────────────────────────────────────────────────────

# Secuencias de mock para dashboard (caché vacío, confirmado_at no existe):
#   fetchone: SHOW_COLUMNS, pedidos_hoy, ventas, producto_top(LIMIT 1), mesas,
#             mesa_top, SHOW_TABLES, categoria_top_row, total_items_row
#   fetchall: estados, estados_hoy
_DASH_FETCHONE_BASE = [
    None,                                                    # SHOW COLUMNS confirmado_at
    {"pedidos_hoy": 5},                                      # pedidos_hoy
    {"ventas_hoy": 1500.0, "ticket_promedio": 300.0},        # ventas + ticket_promedio
    {"nombre": "Smash Burger", "cantidad": 8},               # producto_top (fetchone, LIMIT 1)
    {"mesas_activas": 3},                                    # mesas_activas
    {"numero_mesa": 7, "total_consumido": 950.0},            # mesa_top
    None,                                                    # SHOW TABLES cierres_mesa → no existe
    {"categoria": "Hamburguesas", "cantidad": 8},            # categoria_top_row
    {"total_items": 12},                                     # total_items_row
]
_DASH_FETCHALL_BASE = [
    [{"estado": "listo", "cantidad": 3}],                    # estados (pedidos_activos)
    [{"estado": "listo", "cantidad": 3},
     {"estado": "entregado", "cantidad": 2}],                # estados_hoy
]


class TestDashboardMesaTop:
    def test_mesa_top_presente_en_respuesta(self):
        conn = _mock_conn(_DASH_FETCHONE_BASE, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        assert "mesa_top" in result
        assert result["mesa_top"]["numero"] == 7
        assert abs(result["mesa_top"]["total"] - 950.0) < 0.01

    def test_mesa_top_fallback_cuando_no_hay_ventas(self):
        fetchone = [
            None,
            {"pedidos_hoy": 0},
            {"ventas_hoy": 0.0, "ticket_promedio": 0.0},
            None,                                            # producto_top → None
            {"mesas_activas": 0},
            None,                                            # mesa_top → None (sin ventas)
            None,                                            # SHOW TABLES cierres_mesa
            None,                                            # categoria_top_row → sin ventas
        ]
        conn = _mock_conn(fetchone, [[], []])                # estados vacío, estados_hoy vacío
        result = _call_dashboard(conn)
        assert result["mesa_top"] == {"numero": None, "total": 0.0}
        assert result["categoria_top"] is None
        assert result["tiempo_prep_promedio_min"] is None

    def test_dashboard_incluye_todos_los_campos_requeridos(self):
        conn = _mock_conn(_DASH_FETCHONE_BASE, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        for campo in [
            "ventas_hoy", "pedidos_hoy", "ticket_promedio",
            "pedidos_activos", "estado_pedidos_hoy", "producto_top",
            "categoria_top", "tiempo_prep_promedio_min", "mesa_top",
            "mesas_activas", "cobros_hoy",
        ]:
            assert campo in result, f"Falta campo '{campo}'"

    def test_categoria_top_calcula_porcentaje_sobre_items_del_dia(self):
        conn = _mock_conn(_DASH_FETCHONE_BASE, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        assert result["categoria_top"]["nombre"] == "Hamburguesas"
        assert result["categoria_top"]["cantidad"] == 8
        assert result["categoria_top"]["porcentaje"] == round(8 / 12 * 100, 1)

    def test_estado_pedidos_hoy_incluye_entregado_y_cancelado(self):
        conn = _mock_conn(_DASH_FETCHONE_BASE, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        assert result["estado_pedidos_hoy"]["listo"] == 3
        assert result["estado_pedidos_hoy"]["entregado"] == 2
        assert result["estado_pedidos_hoy"]["cancelado"] == 0

    def test_tiempo_prep_promedio_ausente_sin_columnas_de_trazabilidad(self):
        # _DASH_FETCHONE_BASE simula confirmado_at inexistente (primer None) →
        # pedidos_tiene_columna corta por short-circuit y nunca llega a
        # chequear listo_at ni a calcular el promedio.
        conn = _mock_conn(_DASH_FETCHONE_BASE, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        assert result["tiempo_prep_promedio_min"] is None

    def test_tiempo_prep_promedio_presente_con_columnas_de_trazabilidad(self):
        fetchone = [
            {"Field": "confirmado_at"},                       # SHOW COLUMNS confirmado_at → existe
            {"pedidos_hoy": 5},
            {"ventas_hoy": 1500.0, "ticket_promedio": 300.0},
            {"nombre": "Smash Burger", "cantidad": 8},
            {"mesas_activas": 3},
            {"numero_mesa": 7, "total_consumido": 950.0},
            None,                                              # SHOW TABLES cierres_mesa
            {"categoria": "Hamburguesas", "cantidad": 8},      # categoria_top_row
            {"total_items": 12},                               # total_items_row
            {"Field": "listo_at"},                             # SHOW COLUMNS listo_at → existe
            {"minutos": 17.4},                                 # tiempo_row
        ]
        conn = _mock_conn(fetchone, _DASH_FETCHALL_BASE)
        result = _call_dashboard(conn)
        assert result["tiempo_prep_promedio_min"] == 17.4

    def test_cobros_hoy_incluye_metodos_cuando_cierres_existe(self):
        fetchone = [
            None,
            {"pedidos_hoy": 3},
            {"ventas_hoy": 900.0, "ticket_promedio": 300.0},
            {"nombre": "Papas XL", "cantidad": 6},
            {"mesas_activas": 2},
            {"numero_mesa": 2, "total_consumido": 900.0},
            {"Tables_in_scanorder_db": "cierres_mesa"},      # SHOW TABLES → existe
            {"categoria": "Bebidas", "cantidad": 6},          # categoria_top_row
            {"total_items": 6},                                # total_items_row
        ]
        fetchall = [
            [{"estado": "listo", "cantidad": 3}],
            [
                {"metodo_pago": "efectivo", "cantidad": 2, "total": 600.0},
                {"metodo_pago": "tarjeta",  "cantidad": 1, "total": 300.0},
            ],
            [{"estado": "listo", "cantidad": 3}],              # estados_hoy
        ]
        conn = _mock_conn(fetchone, fetchall)
        result = _call_dashboard(conn)
        assert result["cobros_hoy"]["total"] == 900.0
        assert "efectivo" in result["cobros_hoy"]["metodos"]
        assert "tarjeta"  in result["cobros_hoy"]["metodos"]


# ────────────────────────────────────────────────────────────────────────────
# /reportes/ventas-semana — tendencia de los últimos 7 días
# ────────────────────────────────────────────────────────────────────────────

class TestVentasSemana:
    def test_serie_tiene_siete_dias_incluyendo_hoy(self):
        fetchone = [None]                                     # SHOW COLUMNS confirmado_at
        fetchall = [[]]                                        # sin ventas en el rango
        conn = _mock_conn(fetchone, fetchall)
        result = _call_ventas_semana(conn)
        assert len(result["serie"]) == 7
        assert result["serie"][-1]["fecha"] == date.today().isoformat()
        assert result["serie"][0]["fecha"] == (date.today() - timedelta(days=6)).isoformat()

    def test_dias_sin_ventas_completan_en_cero(self):
        fetchone = [None]
        fetchall = [[]]
        conn = _mock_conn(fetchone, fetchall)
        result = _call_ventas_semana(conn)
        assert all(item["ventas"] == 0.0 and item["pedidos"] == 0 for item in result["serie"])
        assert result["total_ventas"] == 0.0
        assert result["total_pedidos"] == 0

    def test_dia_con_ventas_aparece_en_la_serie(self):
        hoy_iso = date.today().isoformat()
        fetchone = [None]
        fetchall = [[{"fecha": date.today(), "pedidos": 4, "ventas": 800.0}]]
        conn = _mock_conn(fetchone, fetchall)
        result = _call_ventas_semana(conn)
        hoy_item = next(item for item in result["serie"] if item["fecha"] == hoy_iso)
        assert hoy_item["ventas"] == 800.0
        assert hoy_item["pedidos"] == 4
        assert result["total_ventas"] == 800.0


# ────────────────────────────────────────────────────────────────────────────
# /reportes/resumen-hoy — estructura del CSV
# ────────────────────────────────────────────────────────────────────────────

# Secuencias para resumen_hoy_csv (caché vacío):
#   fetchone: SHOW_COLUMNS, resumen, mesa_top, SHOW_TABLES
#   fetchall: por_estado, productos_top, por_hora
_RESUMEN_FETCHONE = [
    None,
    {"ventas_totales": 2400.0, "pedidos_contabilizados": 6, "ticket_promedio": 400.0},
    {"numero_mesa": 5, "total_consumido": 1200.0},
    None,                                                    # cierres_mesa no existe
]
_RESUMEN_FETCHALL = [
    [{"estado": "entregado", "cantidad": 6, "total": 2400.0}],
    [{"nombre": "Smash Burger", "cantidad_vendida": 12, "total_ingresos": 1440.0}],
    [
        {"hora": 12, "pedidos": 3, "ventas": 1200.0},
        {"hora": 19, "pedidos": 3, "ventas": 1200.0},
    ],
]


class TestResumenHoyExport:
    """El endpoint /reportes/resumen-hoy exporta un .xlsx (o .pdf con ?formato=pdf)."""

    def _response(self, fecha=None, formato="excel"):
        conn = _mock_conn(_RESUMEN_FETCHONE, _RESUMEN_FETCHALL)
        return _call_resumen_hoy(conn, fecha=fecha, formato=formato)

    def test_media_type_es_excel(self):
        from app.routes.reportes import EXCEL_MEDIA_TYPE
        assert self._response().media_type == EXCEL_MEDIA_TYPE

    def test_body_es_xlsx_valido(self):
        r = self._response()
        assert r.body[:2] == b"PK", "un .xlsx arranca con la firma ZIP 'PK'"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.body))
        assert wb.worksheets

    def test_content_disposition_incluye_fecha_y_extension(self):
        r = self._response(fecha=date(2026, 3, 20))
        cd = r.headers.get("content-disposition", "")
        assert "2026-03-20" in cd
        assert ".xlsx" in cd

    def test_contiene_secciones_requeridas(self):
        texto = _texto_workbook(self._response())
        for seccion in ["RESUMEN", "PEDIDOS POR ESTADO", "PRODUCTOS MAS VENDIDOS", "VENTAS POR HORA"]:
            assert seccion in texto, f"Falta sección '{seccion}'"

    def test_contiene_metricas_del_resumen(self):
        texto = _texto_workbook(self._response())
        assert "Ventas totales" in texto
        assert "Ticket promedio" in texto

    def test_contiene_mesa_top(self):
        assert "Mesa 5" in _texto_workbook(self._response())

    def test_serie_horaria_tiene_24_horas(self):
        texto = _texto_workbook(self._response())
        for h in range(24):
            assert f"{h:02d}:00" in texto, f"Falta la hora {h:02d}:00"

    def test_acepta_fecha_explicita(self):
        r = self._response(fecha=date(2026, 1, 15))
        assert "2026-01-15" in r.headers.get("content-disposition", "")
        assert "2026-01-15" in _texto_workbook(r)

    def test_sin_cierres_omite_seccion_cobros(self):
        assert "COBROS POR METODO" not in _texto_workbook(self._response())

    def test_con_cierres_incluye_seccion_cobros(self):
        fetchone = [
            None,
            {"ventas_totales": 900.0, "pedidos_contabilizados": 3, "ticket_promedio": 300.0},
            {"numero_mesa": 2, "total_consumido": 900.0},
            {"Tables_in_scanorder_db": "cierres_mesa"},      # existe
        ]
        fetchall = [
            [{"estado": "entregado", "cantidad": 3, "total": 900.0}],
            [{"nombre": "Papas XL", "cantidad_vendida": 9, "total_ingresos": 540.0}],
            [{"metodo_pago": "efectivo", "cantidad": 2, "total": 600.0},
             {"metodo_pago": "tarjeta",  "cantidad": 1, "total": 300.0}],
            [{"hora": 13, "pedidos": 3, "ventas": 900.0}],
        ]
        r = _call_resumen_hoy(_mock_conn(fetchone, fetchall))
        texto = _texto_workbook(r)
        assert "COBROS POR METODO DE PAGO" in texto
        assert "efectivo" in texto
        assert "tarjeta" in texto

    def test_formato_pdf(self):
        from app.routes.reportes import PDF_MEDIA_TYPE
        r = self._response(formato="pdf")
        assert r.media_type == PDF_MEDIA_TYPE
        assert r.body[:4] == b"%PDF"


# ────────────────────────────────────────────────────────────────────────────
# Helpers del módulo (lógica pura sin DB)
# ────────────────────────────────────────────────────────────────────────────

class TestHelpers:
    def test_fecha_venta_sql_usa_created_at_sin_columna(self):
        from app.routes.reportes import fecha_venta_sql
        cursor = MagicMock()
        cursor.fetchone.return_value = None          # confirmado_at no existe
        assert fecha_venta_sql(cursor) == "created_at"

    def test_fecha_venta_sql_usa_coalesce_si_columna_existe(self):
        from app.routes.reportes import fecha_venta_sql
        cursor = MagicMock()
        cursor.fetchone.return_value = {"Field": "confirmado_at"}
        result = fecha_venta_sql(cursor)
        assert "confirmado_at" in result
        assert "COALESCE" in result
