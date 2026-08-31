import io
from datetime import date, timedelta
from html import escape

from fastapi import APIRouter, HTTPException, Response, status, Depends, Query
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.database import close_db_connection, get_db_connection
from app.utils.dependencies import require_role

router = APIRouter(
    prefix="/reportes",
    tags=["Reportes"]
)

ESTADOS_VENTA_SQL = "'confirmado', 'en_preparacion', 'listo', 'entregado'"
_col_cache = {}
EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MEDIA_TYPE = "application/pdf"


def _excel_response(workbook: Workbook, filename: str) -> Response:
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _pdf_response(title: str, subtitle: str, sections: list[dict], filename: str) -> Response:
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=34,
        bottomMargin=34,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ScanOrderTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=22,
        textColor=colors.HexColor("#0B172A"),
        spaceAfter=4,
        alignment=1,
    )
    subtitle_style = ParagraphStyle(
        "ScanOrderSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#52617A"),
        spaceAfter=16,
        alignment=1,
    )
    section_style = ParagraphStyle(
        "ScanOrderSection",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#145F72"),
        spaceBefore=8,
        spaceAfter=6,
    )
    header_cell_style = ParagraphStyle(
        "ScanOrderHeaderCell",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#0B172A"),
    )
    body_cell_style = ParagraphStyle(
        "ScanOrderBodyCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#243044"),
    )

    elements = [Paragraph(title, title_style), Paragraph(subtitle, subtitle_style)]
    for section in sections:
        elements.append(Paragraph(section["title"], section_style))
        headers = section["headers"]
        money_cols = set(section.get("money_cols", ()))
        table_rows = [[Paragraph(escape(str(header)), header_cell_style) for header in headers]]
        for row in section["rows"]:
            table_rows.append([
                Paragraph(
                    escape(_format_money(value) if index in money_cols else str(value)),
                    body_cell_style,
                )
                for index, value in enumerate(row, start=1)
            ])

        col_count = max(len(headers), 1)
        col_widths = {
            2: [250, 210],
            3: [245, 95, 120],
            4: [210, 90, 80, 80],
        }.get(col_count)

        table = Table(table_rows, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F4F7FB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0B172A")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("LEADING", (0, 0), (-1, -1), 11),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E2EC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFCFE")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.extend([table, Spacer(1, 8)])

    doc.build(elements)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type=PDF_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _setup_report_sheet(title: str, subtitle: str):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"
    sheet.append([title])
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    sheet["A1"].font = Font(bold=True, size=16, color="0B172A")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append([subtitle])
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    sheet["A2"].font = Font(size=11, color="52617A")
    sheet["A2"].alignment = Alignment(horizontal="center")
    sheet.append([])
    return workbook, sheet


def _build_excel_report(title: str, subtitle: str, sections: list[dict]) -> Workbook:
    workbook, sheet = _setup_report_sheet(title, subtitle)
    for section in sections:
        _append_section(sheet, section["title"], section["headers"])
        for row in section["rows"]:
            _append_row(sheet, row, money_cols=section.get("money_cols", ()))
    _finish_report_sheet(sheet)
    return workbook


def _append_section(sheet, title: str, headers=None):
    if sheet.max_row > 3:
        sheet.append([])
    sheet.append([title])
    row = sheet.max_row
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    sheet.cell(row=row, column=1).font = Font(bold=True, size=12, color="145F72")
    sheet.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E8F3F7")
    sheet.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    if headers:
        sheet.append(headers)
        header_row = sheet.max_row
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="0B172A")
            cell.fill = PatternFill("solid", fgColor="F4F7FB")
            cell.border = Border(bottom=Side(style="thin", color="D9E2EC"))


def _append_row(sheet, values, money_cols=()):
    sheet.append(values)
    row = sheet.max_row
    for col in money_cols:
        sheet.cell(row=row, column=col).number_format = '$ #,##0.00'


def _format_money(value) -> str:
    number = float(value or 0)
    return f"$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _finish_report_sheet(sheet):
    thin = Side(style="thin", color="E5EAF0")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.value is not None:
                cell.border = Border(bottom=thin)

    widths = {
        "A": 30,
        "B": 22,
        "C": 18,
        "D": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for idx in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(
            sheet.column_dimensions[get_column_letter(idx)].width or 0,
            14,
        )


def pedidos_tiene_columna(cursor, columna: str) -> bool:
    """Indica si la tabla pedidos tiene una columna determinada."""
    key = f"pedidos.{columna}"
    if key not in _col_cache:
        cursor.execute("SHOW COLUMNS FROM pedidos LIKE %s", (columna,))
        _col_cache[key] = cursor.fetchone() is not None
    return _col_cache[key]


def cierres_mesa_existe(cursor) -> bool:
    """Indica si la tabla cierres_mesa ya fue migrada. Resultado cacheado por proceso."""
    return _tabla_existe(cursor, "cierres_mesa")


def _tabla_existe(cursor, tabla: str) -> bool:
    """¿Existe la tabla? Cacheado por proceso. Solo nombres literales del código."""
    key = f"table.{tabla}"
    if key not in _col_cache:
        cursor.execute("SHOW TABLES LIKE %s", (tabla,))
        _col_cache[key] = cursor.fetchone() is not None
    return _col_cache[key]


def _columna_existe(cursor, tabla: str, columna: str) -> bool:
    """¿Existe la columna en la tabla? Cacheado. `tabla` es literal del código."""
    key = f"{tabla}.{columna}"
    if key not in _col_cache:
        cursor.execute(f"SHOW COLUMNS FROM {tabla} LIKE %s", (columna,))
        _col_cache[key] = cursor.fetchone() is not None
    return _col_cache[key]


def fecha_venta_sql(cursor) -> str:
    """
    Usa la fecha de confirmacion de la venta cuando existe.
    Fallback a created_at para bases viejas que todavia no tengan trazabilidad.
    """
    if pedidos_tiene_columna(cursor, "confirmado_at"):
        return "COALESCE(confirmado_at, created_at)"
    return "created_at"


@router.get("/ventas")
def reporte_ventas(
    fecha_inicio: date = Query(..., description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: date    = Query(..., description="Fecha de fin (YYYY-MM-DD)"),
    formato: str = Query("excel", pattern="^(excel|pdf)$", description="Formato de descarga: excel o pdf"),
    current_user: dict = Depends(require_role("admin"))
):
    """
    Exporta un archivo Excel con el reporte de ventas entre las fechas indicadas.
    Incluye resumen, pedidos por estado y productos más vendidos (top 10).
    Solo administradores.
    """
    if fecha_inicio > fecha_fin:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La fecha de inicio no puede ser mayor a la fecha fin"
        )

    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos"
        )
    try:
        cursor = connection.cursor(dictionary=True)
        fecha_venta = fecha_venta_sql(cursor)

        cursor.execute(
            f"""
            SELECT COUNT(*) AS cantidad_pedidos, COALESCE(SUM(total), 0) AS total_ventas
            FROM pedidos
            WHERE estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta}) BETWEEN %s AND %s
            """,
            (fecha_inicio, fecha_fin)
        )
        resumen = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT p.nombre,
                   SUM(dp.cantidad)  AS total_vendido,
                   SUM(dp.subtotal)  AS total_ingresos
            FROM detalle_pedidos dp
            JOIN productos p  ON dp.id_producto = p.id_producto
            JOIN pedidos    pe ON dp.id_pedido   = pe.id_pedido
            WHERE pe.estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta.replace("created_at", "pe.created_at").replace("confirmado_at", "pe.confirmado_at")}) BETWEEN %s AND %s
            GROUP BY p.id_producto, p.nombre
            ORDER BY total_vendido DESC
            LIMIT 10
            """,
            (fecha_inicio, fecha_fin)
        )
        productos = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT estado, COUNT(*) AS cantidad, COALESCE(SUM(total), 0) AS total
            FROM pedidos
            WHERE DATE({fecha_venta}) BETWEEN %s AND %s
            GROUP BY estado
            ORDER BY FIELD(estado, 'pendiente', 'confirmado', 'en_preparacion', 'listo', 'entregado', 'cancelado')
            """,
            (fecha_inicio, fecha_fin)
        )
        estados = cursor.fetchall()

        title = "Maven Burger - Reporte de ventas"
        subtitle = f"Periodo: {fecha_inicio} al {fecha_fin}"
        sections = [
            {
                "title": "RESUMEN",
                "headers": ["Metrica", "Valor"],
                "rows": [
                    ["Total de ventas", _format_money(resumen["total_ventas"])],
                    ["Pedidos contabilizados", resumen["cantidad_pedidos"]],
                    ["Estados de venta", "confirmado, en_preparacion, listo, entregado"],
                ],
            },
            {
                "title": "PEDIDOS POR ESTADO",
                "headers": ["Estado", "Cantidad", "Total"],
                "rows": [
                    [row["estado"] or "sin_estado", row["cantidad"], float(row["total"])]
                    for row in estados
                ],
                "money_cols": (3,),
            },
            {
                "title": "PRODUCTOS MAS VENDIDOS",
                "headers": ["Producto", "Cantidad vendida", "Ingresos"],
                "rows": [
                    [p["nombre"], p["total_vendido"], float(p["total_ingresos"])]
                    for p in productos
                ],
                "money_cols": (3,),
            },
        ]

        extension = formato.lower()
        filename = f"reporte_ventas_{fecha_inicio}_{fecha_fin}.{ 'pdf' if extension == 'pdf' else 'xlsx' }"
        if extension == "pdf":
            return _pdf_response(title, subtitle, sections, filename)
        workbook = _build_excel_report(title, subtitle, sections)
        return _excel_response(workbook, filename)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar reporte: {str(e)}"
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/dashboard")
def dashboard_metricas(current_user: dict = Depends(require_role("admin"))):
    """
    Retorna métricas operativas para el dashboard admin.
    Incluye ventas del día, pedidos activos, ticket promedio, producto destacado y mesas activas.
    """
    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos"
        )
    try:
        cursor = connection.cursor(dictionary=True)
        fecha_venta = fecha_venta_sql(cursor)

        cursor.execute(
            f"""
            SELECT
                COUNT(*) AS pedidos_hoy
            FROM pedidos
            WHERE DATE(created_at) = CURDATE()
            """
        )
        resumen = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT
                COALESCE(SUM(total), 0) AS ventas_hoy,
                COALESCE(AVG(total), 0) AS ticket_promedio
            FROM pedidos
            WHERE estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta}) = CURDATE()
            """
        )
        ventas = cursor.fetchone()

        cursor.execute(
            """
            SELECT estado, COUNT(*) AS cantidad
            FROM pedidos
            WHERE estado IN ('pendiente', 'confirmado', 'en_preparacion', 'listo')
            GROUP BY estado
            """
        )
        estados = {row["estado"]: row["cantidad"] for row in cursor.fetchall()}

        cursor.execute(
            f"""
            SELECT p.nombre, SUM(dp.cantidad) AS cantidad
            FROM detalle_pedidos dp
            JOIN productos p ON p.id_producto = dp.id_producto
            JOIN pedidos pe ON pe.id_pedido = dp.id_pedido
            WHERE DATE({fecha_venta.replace("created_at", "pe.created_at").replace("confirmado_at", "pe.confirmado_at")}) = CURDATE()
            GROUP BY p.id_producto, p.nombre
            ORDER BY cantidad DESC
            LIMIT 1
            """
        )
        producto_top = cursor.fetchone()

        cursor.execute(
            """
            SELECT COUNT(DISTINCT id_mesa) AS mesas_activas
            FROM pedidos
            WHERE estado IN ('pendiente', 'confirmado', 'en_preparacion', 'listo')
            """
        )
        mesas = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT m.numero AS numero_mesa, COALESCE(SUM(p.total), 0) AS total_consumido
            FROM pedidos p
            JOIN mesas m ON m.id_mesa = p.id_mesa
            WHERE p.estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta.replace("created_at", "p.created_at").replace("confirmado_at", "p.confirmado_at")}) = CURDATE()
            GROUP BY p.id_mesa, m.numero
            ORDER BY total_consumido DESC
            LIMIT 1
            """
        )
        mesa_top = cursor.fetchone()

        cobros_hoy = {"total": 0.0, "cantidad": 0, "metodos": {}}
        if cierres_mesa_existe(cursor):
            cursor.execute(
                """
                SELECT
                    COUNT(*) AS cantidad,
                    COALESCE(SUM(total_consumido), 0) AS total,
                    metodo_pago
                FROM cierres_mesa
                WHERE DATE(created_at) = CURDATE()
                GROUP BY metodo_pago
                """
            )
            for row in cursor.fetchall():
                cobros_hoy["cantidad"] += int(row["cantidad"])
                cobros_hoy["total"] += float(row["total"])
                cobros_hoy["metodos"][row["metodo_pago"]] = {
                    "cantidad": int(row["cantidad"]),
                    "total": float(row["total"]),
                }

        # Estado completo de los pedidos de hoy (incluye entregado/cancelado,
        # a diferencia de "pedidos_activos" que solo cubre los en curso).
        cursor.execute(
            """
            SELECT estado, COUNT(*) AS cantidad
            FROM pedidos
            WHERE DATE(created_at) = CURDATE()
            GROUP BY estado
            """
        )
        estados_hoy = {row["estado"]: row["cantidad"] for row in cursor.fetchall()}

        # Categoría más vendida del día (por cantidad de items).
        cursor.execute(
            f"""
            SELECT c.nombre AS categoria, SUM(dp.cantidad) AS cantidad
            FROM detalle_pedidos dp
            JOIN productos p ON p.id_producto = dp.id_producto
            JOIN categorias c ON c.id_categoria = p.id_categoria
            JOIN pedidos pe ON pe.id_pedido = dp.id_pedido
            WHERE DATE({fecha_venta.replace("created_at", "pe.created_at").replace("confirmado_at", "pe.confirmado_at")}) = CURDATE()
              AND pe.estado IN ({ESTADOS_VENTA_SQL})
            GROUP BY c.id_categoria, c.nombre
            ORDER BY cantidad DESC
            LIMIT 1
            """
        )
        categoria_top_row = cursor.fetchone()

        categoria_top = None
        if categoria_top_row:
            cursor.execute(
                f"""
                SELECT COALESCE(SUM(dp.cantidad), 0) AS total_items
                FROM detalle_pedidos dp
                JOIN pedidos pe ON pe.id_pedido = dp.id_pedido
                WHERE DATE({fecha_venta.replace("created_at", "pe.created_at").replace("confirmado_at", "pe.confirmado_at")}) = CURDATE()
                  AND pe.estado IN ({ESTADOS_VENTA_SQL})
                """
            )
            total_items_row = cursor.fetchone()
            total_items = int(total_items_row["total_items"]) if total_items_row else 0
            cantidad_categoria = int(categoria_top_row["cantidad"])
            categoria_top = {
                "nombre": categoria_top_row["categoria"],
                "cantidad": cantidad_categoria,
                "porcentaje": round((cantidad_categoria / total_items) * 100, 1) if total_items else 0.0,
            }

        # Tiempo promedio de preparación (confirmado -> listo). Requiere
        # las columnas de trazabilidad de la migración 004; si no existen,
        # se omite en vez de estimar un número.
        tiempo_prep_promedio_min = None
        if pedidos_tiene_columna(cursor, "confirmado_at") and pedidos_tiene_columna(cursor, "listo_at"):
            cursor.execute(
                """
                SELECT AVG(TIMESTAMPDIFF(MINUTE, confirmado_at, listo_at)) AS minutos
                FROM pedidos
                WHERE DATE(confirmado_at) = CURDATE()
                  AND confirmado_at IS NOT NULL
                  AND listo_at IS NOT NULL
                """
            )
            tiempo_row = cursor.fetchone()
            if tiempo_row and tiempo_row["minutos"] is not None:
                tiempo_prep_promedio_min = round(float(tiempo_row["minutos"]), 1)

        return {
            "ventas_hoy": float(ventas["ventas_hoy"]),
            "pedidos_hoy": int(resumen["pedidos_hoy"]),
            "ticket_promedio": float(ventas["ticket_promedio"]),
            "pedidos_activos": {
                "pendiente": estados.get("pendiente", 0),
                "confirmado": estados.get("confirmado", 0),
                "en_preparacion": estados.get("en_preparacion", 0),
                "listo": estados.get("listo", 0),
            },
            "estado_pedidos_hoy": {
                "pendiente": estados_hoy.get("pendiente", 0),
                "confirmado": estados_hoy.get("confirmado", 0),
                "en_preparacion": estados_hoy.get("en_preparacion", 0),
                "listo": estados_hoy.get("listo", 0),
                "entregado": estados_hoy.get("entregado", 0),
                "cancelado": estados_hoy.get("cancelado", 0),
            },
            "producto_top": producto_top or {"nombre": "—", "cantidad": 0},
            "categoria_top": categoria_top,
            "tiempo_prep_promedio_min": tiempo_prep_promedio_min,
            "mesa_top": (
                {"numero": int(mesa_top["numero_mesa"]), "total": float(mesa_top["total_consumido"])}
                if mesa_top else {"numero": None, "total": 0.0}
            ),
            "mesas_activas": int(mesas["mesas_activas"]),
            "cobros_hoy": cobros_hoy,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener métricas: {str(e)}"
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/resumen-hoy")
def resumen_hoy_excel(
    fecha: date = Query(default=None, description="Fecha (YYYY-MM-DD). Por defecto: hoy."),
    formato: str = Query("excel", pattern="^(excel|pdf)$", description="Formato de descarga: excel o pdf"),
    current_user: dict = Depends(require_role("admin")),
):
    """
    Exporta el resumen operativo del dia como Excel.
    Incluye: ventas totales, cobros por metodo de pago, mesa mas productiva,
    producto top, pedidos por estado y serie horaria.
    """
    fecha_reporte = fecha or date.today()

    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos"
        )
    try:
        cursor = connection.cursor(dictionary=True)
        fecha_venta = fecha_venta_sql(cursor)
        fv_p = fecha_venta.replace("created_at", "p.created_at").replace("confirmado_at", "p.confirmado_at")
        fv_plain = fecha_venta.replace("created_at", "pe.created_at").replace("confirmado_at", "pe.confirmado_at")

        cursor.execute(
            f"""
            SELECT COALESCE(SUM(total), 0) AS ventas_totales,
                   COUNT(*) AS pedidos_contabilizados,
                   COALESCE(AVG(total), 0) AS ticket_promedio
            FROM pedidos
            WHERE estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta}) = %s
            """,
            (fecha_reporte,)
        )
        resumen = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT estado, COUNT(*) AS cantidad, COALESCE(SUM(total), 0) AS total
            FROM pedidos
            WHERE DATE({fecha_venta}) = %s
            GROUP BY estado
            ORDER BY FIELD(estado, 'pendiente','confirmado','en_preparacion','listo','entregado','cancelado')
            """,
            (fecha_reporte,)
        )
        por_estado = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT m.numero AS numero_mesa, COALESCE(SUM(p.total), 0) AS total_consumido
            FROM pedidos p
            JOIN mesas m ON m.id_mesa = p.id_mesa
            WHERE p.estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fv_p}) = %s
            GROUP BY p.id_mesa, m.numero
            ORDER BY total_consumido DESC
            LIMIT 1
            """,
            (fecha_reporte,)
        )
        mesa_top = cursor.fetchone()

        cursor.execute(
            f"""
            SELECT p.nombre, SUM(dp.cantidad) AS cantidad_vendida,
                   SUM(dp.subtotal) AS total_ingresos
            FROM detalle_pedidos dp
            JOIN productos p  ON p.id_producto = dp.id_producto
            JOIN pedidos pe ON pe.id_pedido = dp.id_pedido
            WHERE pe.estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fv_plain}) = %s
            GROUP BY p.id_producto, p.nombre
            ORDER BY cantidad_vendida DESC
            LIMIT 10
            """,
            (fecha_reporte,)
        )
        productos_top = cursor.fetchall()

        cobros_metodos = []
        if cierres_mesa_existe(cursor):
            cursor.execute(
                """
                SELECT metodo_pago,
                       COUNT(*) AS cantidad,
                       COALESCE(SUM(total_consumido), 0) AS total
                FROM cierres_mesa
                WHERE DATE(created_at) = %s
                GROUP BY metodo_pago
                ORDER BY total DESC
                """,
                (fecha_reporte,)
            )
            cobros_metodos = cursor.fetchall()

        cursor.execute(
            f"""
            SELECT HOUR({fecha_venta}) AS hora,
                   COUNT(*) AS pedidos,
                   COALESCE(SUM(total), 0) AS ventas
            FROM pedidos
            WHERE DATE({fecha_venta}) = %s
              AND estado IN ({ESTADOS_VENTA_SQL})
            GROUP BY HOUR({fecha_venta})
            ORDER BY hora
            """,
            (fecha_reporte,)
        )
        por_hora_raw = {int(r["hora"]): r for r in cursor.fetchall()}

        resumen_rows = [
            ["Ventas totales", _format_money(resumen["ventas_totales"])],
            ["Pedidos contabilizados", resumen["pedidos_contabilizados"]],
            ["Ticket promedio", _format_money(resumen["ticket_promedio"])],
        ]
        if mesa_top:
            resumen_rows.append([
                "Mesa mas productiva",
                f"Mesa {mesa_top['numero_mesa']} (${float(mesa_top['total_consumido']):.2f})",
            ])
        else:
            resumen_rows.append(["Mesa mas productiva", "-"])
        if productos_top:
            p = productos_top[0]
            resumen_rows.append(["Producto mas vendido", f"{p['nombre']} ({int(p['cantidad_vendida'])} unidades)"])
        else:
            resumen_rows.append(["Producto mas vendido", "-"])

        sections = [{
            "title": "RESUMEN",
            "headers": ["Metrica", "Valor"],
            "rows": resumen_rows,
        }]
        if cobros_metodos:
            sections.append({
                "title": "COBROS POR METODO DE PAGO",
                "headers": ["Metodo", "Cantidad de cierres", "Total cobrado"],
                "rows": [
                    [row["metodo_pago"], int(row["cantidad"]), float(row["total"])]
                    for row in cobros_metodos
                ],
                "money_cols": (3,),
            })

        sections.extend([
            {
                "title": "PEDIDOS POR ESTADO",
                "headers": ["Estado", "Cantidad", "Total"],
                "rows": [
                    [row["estado"] or "sin_estado", int(row["cantidad"]), float(row["total"])]
                    for row in por_estado
                ],
                "money_cols": (3,),
            },
            {
                "title": "PRODUCTOS MAS VENDIDOS (TOP 10)",
                "headers": ["Producto", "Cantidad vendida", "Ingresos"],
                "rows": [
                    [p["nombre"], int(p["cantidad_vendida"]), float(p["total_ingresos"])]
                    for p in productos_top
                ],
                "money_cols": (3,),
            },
            {
                "title": "VENTAS POR HORA",
                "headers": ["Hora", "Pedidos", "Ventas"],
                "rows": [
                    [
                        f"{hora:02d}:00",
                        int(por_hora_raw[hora]["pedidos"]) if hora in por_hora_raw else 0,
                        float(por_hora_raw[hora]["ventas"]) if hora in por_hora_raw else 0.0,
                    ]
                    for hora in range(24)
                ],
                "money_cols": (3,),
            },
        ])

        title = "Maven Burger - Resumen del dia"
        subtitle = f"Fecha: {fecha_reporte}"
        extension = formato.lower()
        filename = f"resumen_{fecha_reporte}.{ 'pdf' if extension == 'pdf' else 'xlsx' }"
        if extension == "pdf":
            return _pdf_response(title, subtitle, sections, filename)
        workbook = _build_excel_report(title, subtitle, sections)
        return _excel_response(workbook, filename)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar resumen: {str(e)}"
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/ventas-hoy")
def ventas_hoy_por_hora(current_user: dict = Depends(require_role("admin"))):
    """
    Retorna la serie de ventas del dia agrupada por hora para el grafico del dashboard.
    Se contabilizan pedidos confirmados o avanzados en el flujo operativo.
    """
    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos"
        )
    try:
        cursor = connection.cursor(dictionary=True)
        fecha_venta = fecha_venta_sql(cursor)
        cursor.execute(
            f"""
            SELECT
                HOUR({fecha_venta}) AS hora,
                COUNT(*) AS pedidos,
                COALESCE(SUM(total), 0) AS ventas
            FROM pedidos
            WHERE DATE({fecha_venta}) = CURDATE()
              AND estado IN ({ESTADOS_VENTA_SQL})
            GROUP BY HOUR({fecha_venta})
            ORDER BY hora
            """
        )
        ventas_por_hora = {int(row["hora"]): row for row in cursor.fetchall()}

        horas_operativas = list(range(0, 24))
        serie = []
        for hora in horas_operativas:
            row = ventas_por_hora.get(hora)
            serie.append({
                "hora": hora,
                "label": f"{hora:02d}:00",
                "ventas": float(row["ventas"]) if row else 0.0,
                "pedidos": int(row["pedidos"]) if row else 0,
            })

        return {
            "fecha": date.today().isoformat(),
            "total_ventas": sum(item["ventas"] for item in serie),
            "total_pedidos": sum(item["pedidos"] for item in serie),
            "serie": serie,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener ventas por hora: {str(e)}"
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/ventas-semana")
def ventas_ultima_semana(current_user: dict = Depends(require_role("admin"))):
    """
    Retorna la serie de ventas de los ultimos 7 dias (incluye hoy) para el
    grafico de tendencia semanal del dashboard.
    """
    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos"
        )
    try:
        cursor = connection.cursor(dictionary=True)
        fecha_venta = fecha_venta_sql(cursor)
        cursor.execute(
            f"""
            SELECT
                DATE({fecha_venta}) AS fecha,
                COUNT(*) AS pedidos,
                COALESCE(SUM(total), 0) AS ventas
            FROM pedidos
            WHERE estado IN ({ESTADOS_VENTA_SQL})
              AND DATE({fecha_venta}) BETWEEN DATE_SUB(CURDATE(), INTERVAL 6 DAY) AND CURDATE()
            GROUP BY DATE({fecha_venta})
            """
        )
        ventas_por_dia = {row["fecha"].isoformat(): row for row in cursor.fetchall()}

        hoy = date.today()
        serie = []
        for offset in range(6, -1, -1):
            fecha_iso = (hoy - timedelta(days=offset)).isoformat()
            row = ventas_por_dia.get(fecha_iso)
            serie.append({
                "fecha": fecha_iso,
                "ventas": float(row["ventas"]) if row else 0.0,
                "pedidos": int(row["pedidos"]) if row else 0,
            })

        return {
            "total_ventas": sum(item["ventas"] for item in serie),
            "total_pedidos": sum(item["pedidos"] for item in serie),
            "serie": serie,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener ventas de la semana: {str(e)}"
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/mozos")
def reporte_por_mozo(
    fecha_inicio: date = Query(..., description="Fecha de inicio (YYYY-MM-DD)"),
    fecha_fin: date = Query(..., description="Fecha de fin (YYYY-MM-DD)"),
    formato: str = Query("json", pattern="^(json|excel|pdf)$"),
    current_user: dict = Depends(require_role("admin")),
):
    """
    Rendimiento por mozo/usuario en un rango de fechas. Agrega datos que ya se
    guardan, sin auditoría nueva:
      - mesas cerradas + ventas cobradas + ticket promedio  → cierres_mesa.id_usuario_cierre
      - pedidos entregados                                  → movimientos_stock (tipo 'salida').created_by
      - llamados atendidos + tiempo promedio de respuesta   → mozo_llamados.atendido_por
    `formato=json` (default) devuelve la tabla; `excel`/`pdf` la descargan.
    """
    if fecha_inicio > fecha_fin:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La fecha de inicio no puede ser mayor a la fecha fin",
        )

    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos",
        )
    try:
        cursor = connection.cursor(dictionary=True)

        cursor.execute("SELECT id_usuario, nombre, rol, activo FROM usuarios")
        usuarios = {int(u["id_usuario"]): u for u in cursor.fetchall()}

        cierres = {}
        if _tabla_existe(cursor, "cierres_mesa"):
            col_propina = (
                "COALESCE(SUM(propina), 0)"
                if _columna_existe(cursor, "cierres_mesa", "propina")
                else "0"
            )
            cursor.execute(
                f"""
                SELECT id_usuario_cierre AS id_usuario,
                       COUNT(*) AS mesas_cerradas,
                       COALESCE(SUM(total_consumido), 0) AS ventas_cobradas,
                       {col_propina} AS propinas
                FROM cierres_mesa
                WHERE id_usuario_cierre IS NOT NULL
                  AND DATE(created_at) BETWEEN %s AND %s
                GROUP BY id_usuario_cierre
                """,
                (fecha_inicio, fecha_fin),
            )
            cierres = {int(r["id_usuario"]): r for r in cursor.fetchall()}

        entregas = {}
        if _tabla_existe(cursor, "movimientos_stock"):
            cursor.execute(
                """
                SELECT created_by AS id_usuario,
                       COUNT(DISTINCT id_pedido) AS pedidos_entregados
                FROM movimientos_stock
                WHERE tipo = 'salida'
                  AND id_pedido IS NOT NULL
                  AND DATE(created_at) BETWEEN %s AND %s
                GROUP BY created_by
                """,
                (fecha_inicio, fecha_fin),
            )
            entregas = {int(r["id_usuario"]): r for r in cursor.fetchall()}

        llamados = {}
        if _tabla_existe(cursor, "mozo_llamados"):
            cursor.execute(
                """
                SELECT atendido_por AS id_usuario,
                       COUNT(*) AS llamados_atendidos,
                       AVG(TIMESTAMPDIFF(SECOND, solicitado_at, atendido_at)) AS resp_seg
                FROM mozo_llamados
                WHERE atendido_por IS NOT NULL
                  AND atendido_at IS NOT NULL
                  AND DATE(solicitado_at) BETWEEN %s AND %s
                GROUP BY atendido_por
                """,
                (fecha_inicio, fecha_fin),
            )
            llamados = {int(r["id_usuario"]): r for r in cursor.fetchall()}

        # Si en todo el período no hubo ninguna actividad, no listamos el roster
        # entero en cero (se leería como "no hay datos"). Devolvemos lista vacía.
        hay_actividad = bool(cierres or entregas or llamados)
        roster = usuarios if hay_actividad else {}

        filas = []
        for id_usuario, u in roster.items():
            c = cierres.get(id_usuario)
            e = entregas.get(id_usuario)
            ll = llamados.get(id_usuario)
            tiene_actividad = bool(c or e or ll)
            if not u["activo"] and not tiene_actividad:
                continue

            mesas_cerradas = int(c["mesas_cerradas"]) if c else 0
            ventas = float(c["ventas_cobradas"]) if c else 0.0
            propinas = float(c["propinas"]) if c and c.get("propinas") is not None else 0.0
            resp_seg = ll["resp_seg"] if ll and ll["resp_seg"] is not None else None
            filas.append({
                "id_usuario": id_usuario,
                "nombre": u["nombre"],
                "rol": u["rol"],
                "activo": bool(u["activo"]),
                "mesas_cerradas": mesas_cerradas,
                "ventas_cobradas": round(ventas, 2),
                "ticket_promedio": round(ventas / mesas_cerradas, 2) if mesas_cerradas else 0.0,
                "propinas": round(propinas, 2),
                "pedidos_entregados": int(e["pedidos_entregados"]) if e else 0,
                "llamados_atendidos": int(ll["llamados_atendidos"]) if ll else 0,
                "respuesta_promedio_min": round(resp_seg / 60, 1) if resp_seg is not None else None,
            })

        filas.sort(key=lambda f: (-f["ventas_cobradas"], f["nombre"].lower()))

        totales = {
            "mesas_cerradas": sum(f["mesas_cerradas"] for f in filas),
            "ventas_cobradas": round(sum(f["ventas_cobradas"] for f in filas), 2),
            "propinas": round(sum(f["propinas"] for f in filas), 2),
            "pedidos_entregados": sum(f["pedidos_entregados"] for f in filas),
            "llamados_atendidos": sum(f["llamados_atendidos"] for f in filas),
        }

        if formato == "json":
            return {
                "fecha_inicio": fecha_inicio.isoformat(),
                "fecha_fin": fecha_fin.isoformat(),
                "mozos": filas,
                "totales": totales,
            }

        title = "Maven Burger - Reporte por mozo"
        subtitle = f"Periodo: {fecha_inicio} al {fecha_fin}"
        sections = [{
            "title": "RENDIMIENTO POR MOZO",
            "headers": [
                "Mozo", "Rol", "Mesas cerradas", "Ventas cobradas",
                "Ticket prom.", "Propinas", "Pedidos entregados",
                "Llamados atendidos", "Resp. prom. (min)",
            ],
            "rows": [
                [
                    f["nombre"], f["rol"], f["mesas_cerradas"], f["ventas_cobradas"],
                    f["ticket_promedio"], f["propinas"], f["pedidos_entregados"],
                    f["llamados_atendidos"],
                    f["respuesta_promedio_min"] if f["respuesta_promedio_min"] is not None else "-",
                ]
                for f in filas
            ] or [["Sin actividad en el periodo", "", "", "", "", "", "", "", ""]],
            "money_cols": (4, 5, 6),
        }]

        filename = f"reporte_mozos_{fecha_inicio}_{fecha_fin}.{'pdf' if formato == 'pdf' else 'xlsx'}"
        if formato == "pdf":
            return _pdf_response(title, subtitle, sections, filename)
        return _excel_response(_build_excel_report(title, subtitle, sections), filename)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar reporte por mozo: {str(e)}",
        )
    finally:
        cursor.close()
        close_db_connection(connection)


@router.get("/mi-caja")
def mi_caja(
    fecha: date = Query(None, description="Día a consultar (YYYY-MM-DD). Por defecto hoy."),
    current_user: dict = Depends(require_role("mozo", "admin")),
):
    """
    "Mi caja": los cobros que hizo EL usuario actual en un día (por defecto hoy).
    Para que el mozo sepa cuánto efectivo tiene que rendir al cerrar su turno.
    Todo sale de cierres_mesa filtrado por id_usuario_cierre = current_user.
    """
    dia = fecha or date.today()
    id_usuario = current_user.get("user_id")

    vacio = {
        "fecha": dia.isoformat(),
        "resumen": {
            "mesas_cobradas": 0, "total_cobrado": 0.0, "propinas": 0.0,
            "efectivo_a_rendir": 0.0, "por_metodo": {},
        },
        "cobros": [],
    }

    connection = get_db_connection()
    if not connection:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al conectar con la base de datos",
        )
    try:
        cursor = connection.cursor(dictionary=True)
        if not _tabla_existe(cursor, "cierres_mesa"):
            return vacio

        col_prop = "c.propina" if _columna_existe(cursor, "cierres_mesa", "propina") else "0 AS propina"
        cursor.execute(
            f"""
            SELECT c.id_cierre, c.id_mesa, m.numero AS numero_mesa, c.metodo_pago,
                   c.total_consumido, c.vuelto, {col_prop},
                   DATE_FORMAT(c.created_at, '%H:%i') AS hora
            FROM cierres_mesa c
            LEFT JOIN mesas m ON m.id_mesa = c.id_mesa
            WHERE c.id_usuario_cierre = %s
              AND DATE(c.created_at) = %s
            ORDER BY c.created_at ASC
            """,
            (id_usuario, dia),
        )
        rows = cursor.fetchall()

        por_metodo, cobros = {}, []
        total_cobrado = propinas = efectivo_a_rendir = 0.0
        for r in rows:
            tc = float(r["total_consumido"] or 0)
            pr = float(r["propina"] or 0)
            metodo = r["metodo_pago"]
            total_cobrado += tc
            propinas += pr
            por_metodo[metodo] = round(por_metodo.get(metodo, 0.0) + tc, 2)
            # La propina es aparte y es del mozo, no se rinde a la casa: acá
            # va solo la venta cobrada en efectivo.
            if metodo == "efectivo":
                efectivo_a_rendir += tc
            cobros.append({
                "id_cierre": r["id_cierre"],
                "numero_mesa": int(r["numero_mesa"]) if r["numero_mesa"] else None,
                "metodo_pago": metodo,
                "total": round(tc, 2),
                "propina": round(pr, 2),
                "vuelto": round(float(r["vuelto"] or 0), 2),
                "hora": r["hora"],
            })

        return {
            "fecha": dia.isoformat(),
            "resumen": {
                "mesas_cobradas": len(cobros),
                "total_cobrado": round(total_cobrado, 2),
                "propinas": round(propinas, 2),
                "efectivo_a_rendir": round(efectivo_a_rendir, 2),
                "por_metodo": por_metodo,
            },
            "cobros": cobros,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener mi caja: {str(e)}",
        )
    finally:
        cursor.close()
        close_db_connection(connection)
